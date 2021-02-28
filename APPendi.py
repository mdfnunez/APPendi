#excel
import openpyxl
import pandas as pd
import numpy as np


from openpyxl import Workbook
# #APPendi: para la comparacion del registro electronico vs juicio clinico en el diagnostico correcto de apendicitis aguda
import streamlit as st
from PIL import Image
#st.image("https://github.com/mdfnunez/APPendi/blob/main/APPendi.jpg?raw=true")
#st.image("https://github.com/mdfnunez/APPendi/blob/main/Especialidades.jpg?raw=true"
st.image("APPendi.jpg")
st.title("Centro Médico Nacional Siglo XXI")
st.subheader("Protocolo de estudio de pacientes con apendicitis aguda, IMSS")
st.write("")
st.subheader("Registro de nuevo paciente")
Nombre=st.text_input("Nombre")
Edad=st.number_input("Edad")
fechanac=st.date_input("Fecha de nacimiento")
Nss=st.number_input("Numero de seguridad social")
Genero=st.text_input("Genero")


#excel
book=Workbook()
sheet=book.active

sheet["A1"]="Nombre"
sheet["B1"]="Edad"
sheet["C1"]="Fecha de nacimiento"
sheet["D1"]="NSS"
#cambiar color de leta
#sheet["B1"].font=Font(color="FF0000",bold=True)  color rojo
book.save("https://github.com/mdfnunez/APPendi/blob/main/otra%20prueba.xlsx")
#crear otra hoja
#sheet2=book.create_sheet("hoja2")

#leer excel
#book=openpyxl.load_workbook("Nombre de excel para abrir")
#sheet=book.active

#salvar archivos subidos, incompleto
# archivossubidos=st.file_uploader("Subir archivo")
#with open(archivossubidos,"wb") as f:
   # f.write(archivossubidos.getbuffer())
